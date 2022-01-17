import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
#Librerias necesarias

# leemos el fichero
libro = load_workbook('Paciente-2021-12-07.xlsx')

# obtenemos la pestaña/hoja activa (nada mas abrir es la primera)
hoja = libro.active



# Cargamos el archivo de pacientes
df = pd.read_excel("Paciente-2021-12-07.xlsx")

arreglo = np.array([])#Arreglo para almacenar las fechas de registro

size= df.shape[0] #Almacenamos el tamaño del DataFrame
count = 1 #Contador para recorrer los renglones

#Ciclo para recorre los renglones de la columna Fecha y guardarlos
while (count < size ):
    count = count + 1
    arreglo = np.append(arreglo,[hoja.cell(row=count, column=32).value])

#Guardamos las fechas del arreglo en un archivo txt
with open('file.txt', 'w') as temp_file:
    for item in arreglo:
        temp_file.write("%s\n" % item)


with open('file.txt','r') as fechasemin: 
    lineas = [linea.strip() for linea in fechasemin]
#Abrimos el archivo txt y con el metodo strip quitamos los espacios en las fechas


print("Pacientes por mes\n")

#Dentro del archivo txt buscamos las fechas del año 2021
ene21='2021-01'
#Sumamos todas las veces que se repita la fecha
ene21=sum(ene21 in string for string in lineas)
print("Enero: ",ene21)

feb21='2021-02'
feb21=sum(feb21 in string for string in lineas)
print("Febrero: ",feb21)

mar21='2021-03'
mar21=sum(mar21 in string for string in lineas)
print("Marzo: ",mar21)

abr21='2021-04'
abr21=sum(abr21 in string for string in lineas)
print("Abril: ",abr21)

may21='2021-05'
may21=sum(may21 in string for string in lineas)
print("Mayo: ",may21)

jun21='2021-06'
jun21=sum(jun21 in string for string in lineas)
print("Junio: ",jun21)

jul21='2021-07'
jul21=sum(jul21 in string for string in lineas)
print("Julio: ",jul21)

ago21='2021-08'
ago21=sum(ago21 in string for string in lineas)
print("Agosto: ",ago21)

sep21='2021-09'
sep21=sum(sep21 in string for string in lineas)
print("Septiembre: ",sep21)

oct21='2021-10'
oct21=sum(oct21 in string for string in lineas)
print("Octubre: ",oct21)

nov21='2021-11'
nov21=sum(nov21 in string for string in lineas)
print("Noviembre: ",nov21)

dic21 = '2021-12' #lineas son las lineas del file.txt que tienen la fecha
dic21=sum(dic21 in string for string in lineas)
print("Diciembre: ",dic21)

#-------------------------------------------------------------------------

## Declaramos valores para el eje x
eje_x = ['Ene', 'Feb', 'Mar', 'Abr','May', 'Jun', 'Jul', 'Ago','Sep', 'Oct','Nov','Dic']

## Declaramos valores para el eje y
rects1 = [ene21,feb21,mar21,abr21,may21,jun21,jul21,ago21,sep21,oct21,nov21,dic21]

## Creamos Gráfica
plt.bar(eje_x, rects1)
 
## Texto en el eje y
plt.ylabel('Cantidad de usuarios')
 
## Texto  en el eje x
plt.xlabel('Meses')
 
## Título de Gráfica
plt.title('Usuarios en Semin por mes 2021')
 
## Mostramos Gráfica
plt.show()

#Guardamos lq grafica
plt.savefig('Semin por mes 2021.jpg')


