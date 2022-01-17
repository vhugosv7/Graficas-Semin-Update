import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
#Librerias a utilizar 

libro = load_workbook('CitaSucursal-2021-12-07.xlsx')
# obtenemos la pestaña/hoja activa (nada mas abrir es la primera)
hoja = libro.active


#Carga del archivo Excel de citas
df = pd.read_excel("CitaSucursal-2021-12-07.xlsx")
estudios = np.array([])

size= df.shape[0] #Numero total de elementos en el DataFrame 
count2=1

#Bucle para recorrer filas y guardar los estudios 
while (count2 < size ):
    count2 = count2 + 1
    estudios = np.append(estudios,[hoja.cell(row=count2, column=6).value])

#Guardamos todos los tipos de estudios existentes dentro de un archivo txt
with open('fileestudio.txt', 'w') as estudiot:
    for item in estudios:
        estudiot.write("%s\n" % item)

#Abrimos el archivo y guardamos los estudios en una lista
with open('fileestudio.txt','r') as estudiossuma: 
    lineasE = [linea.strip() for linea in estudiossuma]

#Contamos los estudios que hay en el archivo y las veces que se repite
u=dict(zip(lineasE,map(lambda x: lineasE.count(x),lineasE)))

#En un nuevo archivo guardamos el total de veces que se repite cada estudio
with open('archivo.txt', 'w') as archivo:
  archivo.write(str(u))

#Definimos los estudios mas vendidos
Estudios = ['ANTIGENO SARS-CoV-2 (COVID-19 ANTIGENO)',
            'TOMOGRAFIA DE COLUMNA TORACICA MULTICORTE O HELICOIDAL (SIMPLE)',
            'PRUEBA DIAGNOSTICA CONFIRMATORIA SARS-COV-2 (COVID-19)',
            'BIOMETRIA HEMATICA',
            'EXAMEN GENERAL DE ORINA',
            'PRUEBA DIAGNOSTICA CONFIRMATORIA COVID-19',
            'RAYOS X DE TELE DE TORAX PA (1 PROYECCION)',
            'ANTIGENO COVID 19',
            'MASTOGRAFIA',
            'Acido urico, trigliceridos, colesterol total']

#Definimos cada valor de los estudios
valores = [2707,2188,1937,1310,1155,1110,846,687,647,632]

#Asignamos un coor diferente para cada estudio 
colores = ['silver','green', 'yellow','coral','cyan','purple','red','orange','brown','blue']

#Separacion de cada triangulo de la grafica
explode_vals = [0.10,0.10,0.10,0.10,0.10,0.10,0.10,0.10,0.10,0.10]

#Creamos la grafica de pastel con los valores previamente establecidos
plt.pie(x=valores, labels=Estudios, colors = colores, autopct='%1.2f%%', shadow=True, explode = explode_vals)

#Titulo de la grafica
plt.title('Estudios más solicitados 2020-2021')
#Mostramos la grafica 
plt.show()
#Guardamos la grafica
plt.savefig("Mas vendidos.png")
