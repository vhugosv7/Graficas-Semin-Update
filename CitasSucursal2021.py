import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

libro = load_workbook('CitaSucursal-2021-12-07.xlsx')
# obtenemos la pestaña/hoja activa (nada mas abrir es la primera)
hoja = libro.active


df = pd.read_excel("CitaSucursal-2021-12-07.xlsx")#Cargar archivo Citas
arreglo = np.array([])
size= df.shape[0]
count = 1

#Bucle para agregar las fechas de citas en un array
while (count < size ):
    count = count + 1
    arreglo = np.append(arreglo,[hoja.cell(row=count, column=2).value])

#Guardamos las fechas de las citas dentro de una archivo txt
with open('fileCitas.txt', 'w') as citast:
    for item in arreglo:
        citast.write("%s\n" % item)


with open('fileCitas.txt','r') as fcitas: 
    lineas = [linea.strip() for linea in fcitas]

#Buscamos el total de citas
print("Citas en sucursal 2021\n")
ene21='2021-01'
ene21=sum(ene21 in string for string in lineas)
print("Enero: ",ene21)

feb21='2021-02'
feb21=sum(feb21 in string for string in lineas)
print("Febrero: ",feb21)

mar21='2021-03'
mar21=sum(mar21 in string for string in lineas)
print("Marzo: ",mar21)

abr21='2021-04'
abr21=sum(abr21 in string for string in lineas)
print("Abril: ",abr21)

may21='2021-05'
may21=sum(may21 in string for string in lineas)
print("Mayo: ",may21)

jun21='2021-06'
jun21=sum(jun21 in string for string in lineas)
print("Junio: ",jun21)

jul21='2021-07'
jul21=sum(jul21 in string for string in lineas)
print("Julio: ",jul21)

ago21='2021-08'
ago21=sum(ago21 in string for string in lineas)
print("Agosto: ",ago21)

sep21='2021-09'
sep21=sum(sep21 in string for string in lineas)
print("Septiembre: ",sep21)

oct21='2021-10'
oct21=sum(oct21 in string for string in lineas)
print("Octubre: ",oct21)

nov21='2021-11'
nov21=sum(nov21 in string for string in lineas)
print("Noviembre: ",nov21)

dic21='2021-12'
dic21=sum(dic21 in string for string in lineas)
print("Diciembre: ",dic21)


## Declaramos valores para el eje x
eje_x = ['Ene', 'Feb', 'Mar', 'Abr','May', 'Jun', 'Jul', 'Ago','Sep', 'Oct','Nov','Dic']


## Declaramos valores para el eje y
rects1 = [ene21,feb21,mar21,abr21,may21,jun21,jul21,ago21,sep21,oct21,nov21,dic21]

## Creamos Gráfica
plt.bar(eje_x, rects1, color="green")
 
##Texto en el eje y
plt.ylabel('Cantidad ')
 
## Texto en el eje x
plt.xlabel('Meses')
 
## Título de Gráfica
plt.title('Citas en sucursal 2021')
 
## Mostramos Gráfica
plt.show()

#Guardamos la grafica
plt.savefig("Citas en sucursal 2021.png", bbox_inches='tight')
